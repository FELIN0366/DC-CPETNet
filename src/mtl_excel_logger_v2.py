#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
MTL Excel Logger V2 - 多任务学习评估结果落盘模块（19 Sheet 设计）
================================================================

功能：
- 每个 fold 创建 3 个 sheet：
  1. val_fold{N}: 该 fold 的 val 结果（使用最佳阈值）
  2. holdout_fold{N}: 该 fold 的 holdout 结果（使用最佳阈值）
  3. threshold_fold{N}: 该 fold 的最优阈值
- 最终汇总 4 个 sheet：
  1. val_summary: 五折汇总（mean/std）
  2. holdout_summary: holdout 汇总（mean/std）
  3. threshold_summary: 阈值汇总（mean/std）
  4. config_snapshot: config.yaml 快照

总共 19 个 sheet（5×3 + 4）

关键设计：
- config_snapshot 在实验开始时立即写入并保存
- 每个 fold 完成后立即写入 3 个 sheet 并保存（防止崩溃丢失）
- 五折完成后追加汇总 4 个 sheet

创建日期: 2026-04-22
更新日期: 2026-04-22 (V2 - 19 sheet 设计)
"""

import os
import statistics
from typing import Dict, Any, Optional, List
from datetime import datetime
import json

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[Warning] openpyxl 未安装，Excel 功能不可用")


class MTLExcelLoggerV2:
    """
    MTL Excel Logger V2 - 19 Sheet 设计

    Sheet 结构：
    ============

    【每个 Fold 的 Sheet（5×3 = 15 个）】
    - val_fold1 ~ val_fold5: 各 fold 的验证集结果
    - holdout_fold1 ~ holdout_fold5: 各 fold 的 holdout 测试集结果
    - threshold_fold1 ~ threshold_fold5: 各 fold 的最优阈值

    【汇总 Sheet（4 个）】
    - val_summary: 五折验证集汇总（mean/std/min/max）
    - holdout_summary: holdout 测试集汇总（mean/std/min/max）
    - threshold_summary: 阈值汇总（mean/std）
    - config_snapshot: config.yaml 快照

    列结构（val/holdout sheet）：
    - task_key: 任务键 (t1~t6)
    - task_name: 任务名称
    - acc: 准确率
    - precision: 精确率
    - recall: 召回率
    - f1: F1分数
    - macro_f1: Macro F1分数
    - auc: AUC (仅二分类任务 t3/t4/t5)
    - threshold: 使用的阈值 (仅二分类任务)
    - minority_f1: 少数类F1 (仅二分类)
    - minority_recall: 少数类召回率 (仅二分类)
    - minority_precision: 少数类精确率 (仅二分类)
    - pred_minor_rate: 预测少数类比例 (仅二分类)
    - true_minor_rate: 真实少数类比例 (仅二分类)
    - minority_tp: 少数类TP (仅二分类)
    - minority_fn: 少数类FN (仅二分类)
    - majority_fp: 多数类FP (仅二分类)
    - majority_tn: 多数类TN (仅二分类)
    - notes: 备注

    列结构（summary sheet）：
    - task_key, task_name, metric_name, MEAN, STD, MIN, MAX, notes
    - 每个任务的每个指标各一行 (macro_f1 所有任务; minority_*/auc 仅二分类)

    列结构（threshold sheet）：
    - task_key: 任务键 (t3/t4/t5)
    - task_name: 任务名称
    - best_threshold: 最佳阈值
    - best_f1: 最佳F1分数
    - baseline_f1: 基线F1分数(阈值0.5)
    - improvement: 提升幅度
    """

    # val/holdout sheet 列名
    METRIC_COLUMNS = [
        "task_key", "task_name", "acc", "precision",
        "recall", "f1", "macro_f1", "auc", "auprc", "threshold",
        # 二分类任务扩展指标 (t3/t4/t5)
        "minority_f1", "minority_recall", "minority_precision",
        "pred_minor_rate", "true_minor_rate",
        "minority_tp", "minority_fn", "majority_fp", "majority_tn",
        "notes"
    ]

    # threshold sheet 列名
    THRESHOLD_COLUMNS = [
        "task_key", "task_name", "best_threshold",
        "best_f1", "baseline_f1", "improvement"
    ]

    # 二分类任务列表
    BINARY_TASKS = ["t3", "t4", "t5"]

    def __init__(
        self,
        excel_path: str,
        checkpoint_name: str = "mtl_v31_t6_protected",
        config_dict: Optional[Dict[str, Any]] = None,
        raw_yaml_content: Optional[str] = None,
        n_folds: int = 5,
        force_new: bool = True
    ):
        """
        初始化 Excel Logger V2

        Args:
            excel_path: Excel 文件路径
            checkpoint_name: checkpoint 名称 (用于标识)
            config_dict: 配置字典 (用于基本信息和摘要)
            raw_yaml_content: 完整的 YAML 文件原始文本内容 (优先使用，直接保存)
            n_folds: Fold 总数 (默认 5)
            force_new: 是否强制创建新文件 (默认 True)
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl 未安装，请先安装: pip install openpyxl")

        self.excel_path = excel_path
        self.checkpoint_name = checkpoint_name
        self.config_dict = config_dict or {}
        self.raw_yaml_content = raw_yaml_content
        self.n_folds = n_folds
        self.completed_folds = 0

        # 缓存各 fold 数据（用于最终汇总）
        self.fold_val_data: Dict[int, Dict[str, Dict[str, float]]] = {}
        self.fold_holdout_data: Dict[int, Dict[str, Dict[str, float]]] = {}
        self.fold_threshold_data: Dict[int, Dict[str, Dict[str, float]]] = {}

        # 样式定义
        self._setup_styles()

        # 初始化 Excel 文件
        self._init_excel(force_new)

        # 立即写入 config_snapshot 并保存
        if config_dict or raw_yaml_content:
            self.write_config_snapshot(config_dict or {}, raw_yaml_content)
            self.save()
            if raw_yaml_content:
                print(f"[MTLExcelLoggerV2] config_snapshot 已写入完整 YAML 内容并保存")
            else:
                print(f"[MTLExcelLoggerV2] config_snapshot 已写入并保存 (字典格式)")

    def _setup_styles(self):
        """设置 Excel 样式"""
        # 表头样式
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        # 普通单元格样式
        self.cell_alignment = Alignment(horizontal="center", vertical="center")
        self.number_format = "0.0000"
        self.threshold_format = "0.000"

        # 汇总行样式
        self.summary_font = Font(bold=True, size=11, color="0000FF")
        self.summary_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

        # 边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.cell_border = thin_border

    def _init_excel(self, force_new: bool):
        """初始化 Excel 文件"""
        # 确保目录存在
        os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)

        if force_new or not os.path.exists(self.excel_path):
            # 创建新文件
            self.wb = Workbook()

            # 删除默认 sheet
            if "Sheet" in self.wb.sheetnames:
                del self.wb["Sheet"]

            # 创建汇总 sheet（初始为空）
            self._create_summary_sheet("val_summary")
            self._create_summary_sheet("holdout_summary")
            self._create_threshold_summary_sheet("threshold_summary")
            self._create_config_sheet("config_snapshot")

            self.wb.save(self.excel_path)
            print(f"[MTLExcelLoggerV2] 创建新文件: {self.excel_path}")
        else:
            # 加载已有文件
            self.wb = openpyxl.load_workbook(self.excel_path)
            print(f"[MTLExcelLoggerV2] 加载已有文件: {self.excel_path}")

    def _create_fold_sheet(self, sheet_name: str, is_threshold: bool = False):
        """创建 fold sheet 并写入表头"""
        ws = self.wb.create_sheet(title=sheet_name)

        if is_threshold:
            # threshold sheet 表头
            columns = self.THRESHOLD_COLUMNS
            ws.column_dimensions['A'].width = 10   # task_key
            ws.column_dimensions['B'].width = 20   # task_name
            ws.column_dimensions['C'].width = 15   # best_threshold
            ws.column_dimensions['D'].width = 12   # best_f1
            ws.column_dimensions['E'].width = 12   # baseline_f1
            ws.column_dimensions['F'].width = 12   # improvement
        else:
            # val/holdout sheet 表头
            columns = self.METRIC_COLUMNS
            ws.column_dimensions['A'].width = 10   # task_key
            ws.column_dimensions['B'].width = 20   # task_name
            ws.column_dimensions['C'].width = 10   # acc
            ws.column_dimensions['D'].width = 12   # precision
            ws.column_dimensions['E'].width = 10   # recall
            ws.column_dimensions['F'].width = 10   # f1
            ws.column_dimensions['G'].width = 12   # macro_f1
            ws.column_dimensions['H'].width = 10   # auc
            ws.column_dimensions['I'].width = 12   # threshold
            ws.column_dimensions['J'].width = 12   # minority_f1
            ws.column_dimensions['K'].width = 14   # minority_recall
            ws.column_dimensions['L'].width = 16   # minority_precision
            ws.column_dimensions['M'].width = 16   # pred_minor_rate
            ws.column_dimensions['N'].width = 16   # true_minor_rate
            ws.column_dimensions['O'].width = 12   # minority_tp
            ws.column_dimensions['P'].width = 12   # minority_fn
            ws.column_dimensions['Q'].width = 12   # majority_fp
            ws.column_dimensions['R'].width = 12   # majority_tn
            ws.column_dimensions['S'].width = 30   # notes

        # 写入表头
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.cell_border

    # 汇总 sheet 的指标行定义
    SUMMARY_METRIC_ROWS = [
        ("macro_f1", "macro_f1"),
        ("minority_f1", "minority_f1 (二分类)"),
        ("minority_recall", "minority_recall (二分类)"),
        ("minority_precision", "minority_precision (二分类)"),
        ("auc", "auc (二分类)"),
    ]

    def _create_summary_sheet(self, sheet_name: str):
        """创建汇总 sheet（含统计列）"""
        ws = self.wb.create_sheet(title=sheet_name)

        # 汇总列名: task_key, task_name, metric_name, MEAN, STD, MIN, MAX, notes
        columns = ["task_key", "task_name", "metric_name", "MEAN", "STD", "MIN", "MAX", "notes"]

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.cell_border

        # 列宽设置
        ws.column_dimensions['A'].width = 10   # task_key
        ws.column_dimensions['B'].width = 20   # task_name
        ws.column_dimensions['C'].width = 25   # metric_name
        ws.column_dimensions['D'].width = 12   # MEAN
        ws.column_dimensions['E'].width = 12   # STD
        ws.column_dimensions['F'].width = 12   # MIN
        ws.column_dimensions['G'].width = 12   # MAX
        ws.column_dimensions['H'].width = 30   # notes

    def _create_threshold_summary_sheet(self, sheet_name: str):
        """创建阈值汇总 sheet"""
        ws = self.wb.create_sheet(title=sheet_name)

        # 阈值汇总列名
        columns = ["task_key", "task_name", "MEAN", "STD", "MIN", "MAX", "notes"]

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.cell_border

        # 列宽设置
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        for i in range(3, 8):
            ws.column_dimensions[chr(64 + i)].width = 12

    def _create_config_sheet(self, sheet_name: str):
        """创建配置 sheet"""
        ws = self.wb.create_sheet(title=sheet_name)
        ws.column_dimensions['A'].width = 25  # config_key
        ws.column_dimensions['B'].width = 80  # config_value

    def write_config_snapshot(self, config_dict: Dict[str, Any], raw_yaml_content: Optional[str] = None):
        """
        写入配置快照

        Args:
            config_dict: 配置字典（用于基本信息和摘要）
            raw_yaml_content: 完整的 YAML 文件原始文本内容（优先使用）
        """
        ws = self.wb["config_snapshot"]

        # 清空已有内容
        ws.delete_rows(1, ws.max_row)

        # 如果提供了原始 YAML 内容，直接保存完整文件
        if raw_yaml_content:
            # 写入基本信息作为表头
            config_items = [
                ("=== 基本信息 ===", ""),
                ("checkpoint_name", self.checkpoint_name),
                ("created_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                ("excel_path", self.excel_path),
                ("n_folds", self.n_folds),
                ("config_file", "configs/config_mtl.yaml (完整内容如下)"),
                ("", ""),
                ("=== config_mtl.yaml 完整内容 ===", ""),
            ]

            # 写入表头
            for row_idx, (key, value) in enumerate(config_items, start=1):
                ws.cell(row=row_idx, column=1, value=key)
                ws.cell(row=row_idx, column=2, value=value)

            # 写入完整 YAML 内容（逐行写入，每行一个单元格）
            yaml_lines = raw_yaml_content.split('\n')
            start_row = len(config_items) + 1
            for line_idx, line in enumerate(yaml_lines):
                ws.cell(row=start_row + line_idx, column=1, value=line)

            # 调整列宽以显示完整 YAML 内容
            ws.column_dimensions['A'].width = 120  # 增大宽度以容纳 YAML 行
        else:
            # 回退到原有逻辑：按字典格式写入
            config_items = [
                ("=== 基本信息 ===", ""),
                ("checkpoint_name", self.checkpoint_name),
                ("created_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                ("excel_path", self.excel_path),
                ("n_folds", self.n_folds),
                ("", ""),
            ]

            # 添加任务配置
            if "tasks" in config_dict:
                config_items.append(("=== 任务配置 ===", ""))
                tasks_config = config_dict["tasks"]
                for task_key in sorted(tasks_config.keys()):
                    task_cfg = tasks_config[task_key]
                    if isinstance(task_cfg, dict):
                        config_items.append(
                            (f"task.{task_key}", json.dumps(task_cfg, ensure_ascii=False, indent=2))
                        )
                    else:
                        config_items.append((f"task.{task_key}", str(task_cfg)))
                config_items.append(("", ""))

            # 添加训练配置
            if "training" in config_dict:
                config_items.append(("=== 训练配置 ===", ""))
                training_config = config_dict["training"]
                for key, value in training_config.items():
                    if isinstance(value, dict):
                        config_items.append((f"training.{key}", json.dumps(value, ensure_ascii=False)))
                    else:
                        config_items.append((f"training.{key}", str(value)))
                config_items.append(("", ""))

            # 添加数据配置
            if "data" in config_dict:
                config_items.append(("=== 数据配置 ===", ""))
                data_config = config_dict["data"]
                for key, value in data_config.items():
                    config_items.append((f"data.{key}", str(value)))
                config_items.append(("", ""))

            # 添加 MTL 配置
            if "mtl" in config_dict:
                config_items.append(("=== MTL 配置 ===", ""))
                mtl_cfg = config_dict["mtl"]
                for key, value in mtl_cfg.items():
                    if isinstance(value, dict):
                        config_items.append((f"mtl.{key}", json.dumps(value, ensure_ascii=False, indent=2)))
                    else:
                        config_items.append((f"mtl.{key}", str(value)))

            # 写入到 sheet
            for row_idx, (key, value) in enumerate(config_items, start=1):
                ws.cell(row=row_idx, column=1, value=key)
                ws.cell(row=row_idx, column=2, value=value)

    def write_fold_results(
        self,
        fold: int,
        val_metrics: Dict[str, Dict[str, float]],
        holdout_metrics: Dict[str, Dict[str, float]],
        thresholds: Dict[str, Dict[str, float]],
        task_names: Optional[Dict[str, str]] = None
    ):
        """
        写入单个 fold 的全部结果（3 个 sheet）

        Args:
            fold: Fold 编号 (1-5)
            val_metrics: 验证集指标（使用最佳阈值）
                {"t1": {"acc": 0.85, "f1": 0.82, ...}, "t3": {...}, ...}
            holdout_metrics: Holdout 测试集指标（使用最佳阈值）
                {"t1": {"acc": 0.83, ...}, ...}
            thresholds: 阈值搜索结果
                {"t3": {"best_threshold": 0.325, "best_f1": 0.72, "baseline_f1": 0.68, "improvement": 0.04}, ...}
            task_names: 任务名称字典 {"t1": "运动心功能分级", ...}
        """
        # 1. 写入 val_fold{N}
        self._write_metric_sheet(
            f"val_fold{fold}",
            val_metrics,
            thresholds,
            task_names,
            f"Fold {fold} 验证集结果（使用最佳阈值）"
        )

        # 2. 写入 holdout_fold{N}
        self._write_metric_sheet(
            f"holdout_fold{fold}",
            holdout_metrics,
            thresholds,
            task_names,
            f"Fold {fold} Holdout 测试集结果（使用最佳阈值）"
        )

        # 3. 写入 threshold_fold{N}
        self._write_threshold_sheet(f"threshold_fold{fold}", thresholds, task_names)

        # 缓存数据用于汇总
        self.fold_val_data[fold] = val_metrics
        self.fold_holdout_data[fold] = holdout_metrics
        self.fold_threshold_data[fold] = thresholds

        # 立即保存
        self.save()
        self.completed_folds += 1
        print(f"[MTLExcelLoggerV2] Fold {fold} 完成：已写入 3 个 sheet 并保存")

    # 二分类任务扩展指标键名
    BINARY_EXTRA_KEYS = [
        "minority_f1", "minority_recall", "minority_precision",
        "pred_minor_rate", "true_minor_rate",
        "minority_tp", "minority_fn", "majority_fp", "majority_tn"
    ]

    def _write_metric_sheet(
        self,
        sheet_name: str,
        metrics: Dict[str, Dict[str, float]],
        thresholds: Dict[str, Dict[str, float]],
        task_names: Optional[Dict[str, str]],
        notes: str
    ):
        """写入指标 sheet（val 或 holdout）"""
        # 创建 sheet
        self._create_fold_sheet(sheet_name, is_threshold=False)
        ws = self.wb[sheet_name]

        # 写入每个任务的指标
        row_idx = 2
        for task_key in sorted(metrics.keys()):
            task_metrics = metrics[task_key]
            task_name = task_names.get(task_key, task_key) if task_names else task_key
            is_binary = task_key in self.BINARY_TASKS

            # 获取阈值（仅二分类任务）
            threshold_val = "N/A"
            if is_binary and task_key in thresholds:
                threshold_val = thresholds[task_key].get("best_threshold", 0.5)

            # 构建数据行
            row_data = [
                task_key,                                                          # A: task_key
                task_name,                                                         # B: task_name
                task_metrics.get("acc", task_metrics.get("accuracy", 0)),          # C: acc
                task_metrics.get("precision", 0),                                  # D: precision
                task_metrics.get("recall", 0),                                     # E: recall
                task_metrics.get("f1", 0),                                         # F: f1
                task_metrics.get("macro_f1", task_metrics.get("f1", 0)),           # G: macro_f1
                task_metrics.get("auc", 0) if is_binary else "N/A",               # H: auc
                task_metrics.get("auprc", 0) if is_binary else "N/A",             # I: auprc
                threshold_val,                                                     # J: threshold
            ]

            # 二分类任务扩展指标 (K-S)
            for key in self.BINARY_EXTRA_KEYS:
                if is_binary:
                    row_data.append(task_metrics.get(key, "N/A"))
                else:
                    row_data.append("N/A")

            # notes (T)
            row_data.append(notes if row_idx == 2 else "")

            # 写入单元格
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = self.cell_alignment
                cell.border = self.cell_border

                # 数值格式化
                if isinstance(value, float):
                    if col_idx == 10:  # threshold 列 (J)
                        cell.number_format = self.threshold_format
                    elif col_idx >= 3:  # 数值列
                        cell.number_format = self.number_format

            row_idx += 1

    def _write_threshold_sheet(
        self,
        sheet_name: str,
        thresholds: Dict[str, Dict[str, float]],
        task_names: Optional[Dict[str, str]]
    ):
        """写入阈值 sheet"""
        # 创建 sheet
        self._create_fold_sheet(sheet_name, is_threshold=True)
        ws = self.wb[sheet_name]

        # 写入每个二分类任务的阈值信息
        row_idx = 2
        for task_key in self.BINARY_TASKS:
            if task_key not in thresholds:
                continue

            task_threshold = thresholds[task_key]
            task_name = task_names.get(task_key, task_key) if task_names else task_key

            # 构建数据行
            row_data = [
                task_key,
                task_name,
                task_threshold.get("best_threshold", 0.5),
                task_threshold.get("best_f1", 0),
                task_threshold.get("baseline_f1", 0),
                task_threshold.get("improvement", 0)
            ]

            # 写入单元格
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = self.cell_alignment
                cell.border = self.cell_border

                # 数值格式化
                if isinstance(value, float):
                    if col_idx == 3:  # threshold 列
                        cell.number_format = self.threshold_format
                    else:
                        cell.number_format = self.number_format

            row_idx += 1

    def write_summary_sheets(self, task_names: Optional[Dict[str, str]] = None):
        """
        写入汇总 sheet（五折完成后调用）

        Args:
            task_names: 任务名称字典
        """
        if self.completed_folds < self.n_folds:
            print(f"[Warning] 仅有 {self.completed_folds}/{self.n_folds} 个 fold 完成，汇总可能不完整")

        # 1. val_summary
        self._write_metric_summary_sheet("val_summary", self.fold_val_data, task_names)

        # 2. holdout_summary
        self._write_metric_summary_sheet("holdout_summary", self.fold_holdout_data, task_names)

        # 3. threshold_summary
        self._write_threshold_summary_sheet("threshold_summary", task_names)

        # 立即保存
        self.save()
        print(f"[MTLExcelLoggerV2] 汇总完成：已写入 4 个汇总 sheet 并保存")

    # 汇总 sheet 中要统计的指标键名（除 macro_f1 外的通用指标）
    SUMMARY_BASIC_KEYS = ["acc", "accuracy", "precision", "recall", "f1", "macro_f1", "auc", "auprc"]
    # 二分类扩展指标键名
    SUMMARY_BINARY_KEYS = [
        "minority_f1", "minority_recall", "minority_precision",
        "pred_minor_rate", "true_minor_rate"
    ]

    def _write_metric_summary_sheet(
        self,
        sheet_name: str,
        fold_data: Dict[int, Dict[str, Dict[str, float]]],
        task_names: Optional[Dict[str, str]]
    ):
        """写入指标汇总 sheet（每个任务的每个指标各一行）"""
        ws = self.wb[sheet_name]

        # 清空已有内容（保留表头）
        ws.delete_rows(2, ws.max_row)

        # 获取所有任务键
        all_task_keys = set()
        for fold_metrics in fold_data.values():
            all_task_keys.update(fold_metrics.keys())

        # 每个任务要统计的指标行
        row_idx = 2
        for task_key in sorted(all_task_keys):
            task_name = task_names.get(task_key, task_key) if task_names else task_key
            is_binary = task_key in self.BINARY_TASKS

            # 确定该任务要统计的指标
            metric_keys = ["macro_f1"]  # 总是统计 macro_f1
            if is_binary:
                metric_keys.extend(self.SUMMARY_BINARY_KEYS)
                metric_keys.append("auc")
                metric_keys.append("auprc")

            for metric_key in metric_keys:
                # 收集该任务所有 fold 的该指标值
                values = []
                for fold in fold_data:
                    if task_key in fold_data[fold]:
                        val = fold_data[fold][task_key].get(metric_key)
                        if isinstance(val, (int, float)):
                            values.append(val)

                if not values:
                    continue

                mean_val = statistics.mean(values)
                std_val = statistics.stdev(values) if len(values) > 1 else 0
                min_val = min(values)
                max_val = max(values)

                # 显示名称
                display_name = metric_key
                for mk, dn in self.SUMMARY_METRIC_ROWS:
                    if mk == metric_key:
                        display_name = dn
                        break

                row_data = [
                    task_key,
                    task_name,
                    display_name,
                    mean_val,
                    std_val,
                    min_val,
                    max_val,
                    f"基于 {len(values)} 个 fold"
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = self.cell_alignment
                    cell.border = self.cell_border
                    cell.font = self.summary_font
                    cell.fill = self.summary_fill

                    if isinstance(value, float) and col_idx >= 4:
                        cell.number_format = self.number_format

                row_idx += 1

    def _write_threshold_summary_sheet(
        self,
        sheet_name: str,
        task_names: Optional[Dict[str, str]]
    ):
        """写入阈值汇总 sheet"""
        ws = self.wb[sheet_name]

        # 清空已有内容（保留表头）
        ws.delete_rows(2, ws.max_row)

        # 按二分类任务计算阈值统计
        row_idx = 2
        for task_key in self.BINARY_TASKS:
            task_name = task_names.get(task_key, task_key) if task_names else task_key

            # 收集该任务所有 fold 的阈值
            threshold_values = []
            for fold in self.fold_threshold_data:
                if task_key in self.fold_threshold_data[fold]:
                    thresh = self.fold_threshold_data[fold][task_key].get("best_threshold", 0.5)
                    threshold_values.append(thresh)

            if threshold_values:
                mean_val = statistics.mean(threshold_values)
                std_val = statistics.stdev(threshold_values) if len(threshold_values) > 1 else 0
                min_val = min(threshold_values)
                max_val = max(threshold_values)
            else:
                mean_val, std_val, min_val, max_val = 0.5, 0, 0.5, 0.5

            # 构建汇总行
            row_data = [
                task_key,
                task_name,
                mean_val,
                std_val,
                min_val,
                max_val,
                f"基于 {len(threshold_values)} 个 fold"
            ]

            # 写入单元格
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = self.cell_alignment
                cell.border = self.cell_border
                cell.font = self.summary_font
                cell.fill = self.summary_fill

                if isinstance(value, float) and col_idx >= 3:
                    cell.number_format = self.threshold_format

            row_idx += 1

    def load_existing_fold_data(self):
        """
        从已有的 Excel sheet 中加载已完成 fold 的数据到内存缓存。
        用于 resume 场景：start_fold > 1 时，前面的 fold 数据已在 Excel 中，
        需要读回到 fold_val_data / fold_holdout_data / fold_threshold_data 以便最终汇总。
        """
        loaded_folds = []
        for fold in range(1, self.n_folds + 1):
            val_sheet = f"val_fold{fold}"
            holdout_sheet = f"holdout_fold{fold}"
            threshold_sheet = f"threshold_fold{fold}"

            if val_sheet not in self.wb.sheetnames:
                continue

            # 读取 val sheet
            val_data = {}
            ws = self.wb[val_sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                task_key = str(row[0])
                val_data[task_key] = {}
                for i, col_name in enumerate(self.METRIC_COLUMNS):
                    if i == 0 or i >= len(row):
                        continue
                    val = row[i] if i < len(row) else None
                    if val is not None and val != "N/A":
                        val_data[task_key][col_name] = val

            # 读取 holdout sheet
            holdout_data = {}
            if holdout_sheet in self.wb.sheetnames:
                ws = self.wb[holdout_sheet]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        continue
                    task_key = str(row[0])
                    holdout_data[task_key] = {}
                    for i, col_name in enumerate(self.METRIC_COLUMNS):
                        if i == 0 or i >= len(row):
                            continue
                        val = row[i] if i < len(row) else None
                        if val is not None and val != "N/A":
                            holdout_data[task_key][col_name] = val

            # 读取 threshold sheet
            threshold_data = {}
            if threshold_sheet in self.wb.sheetnames:
                ws = self.wb[threshold_sheet]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        continue
                    task_key = str(row[0])
                    threshold_data[task_key] = {}
                    for i, col_name in enumerate(self.THRESHOLD_COLUMNS):
                        if i == 0 or i >= len(row):
                            continue
                        val = row[i] if i < len(row) else None
                        if val is not None:
                            threshold_data[task_key][col_name] = val

            # 只有数据非空才记录
            if val_data:
                self.fold_val_data[fold] = val_data
                self.fold_holdout_data[fold] = holdout_data
                self.fold_threshold_data[fold] = threshold_data
                loaded_folds.append(fold)

        self.completed_folds = len(loaded_folds)
        if loaded_folds:
            print(f"[MTLExcelLoggerV2] 已从 Excel 加载 {len(loaded_folds)} 个 fold 的数据: folds {loaded_folds}")
        return loaded_folds

    def save(self):
        """强制保存 Excel 文件"""
        self.wb.save(self.excel_path)


def create_mtl_excel_logger_v2(
    excel_path: str,
    checkpoint_name: str = "mtl_v31_t6_protected",
    config_dict: Optional[Dict[str, Any]] = None,
    raw_yaml_content: Optional[str] = None,
    n_folds: int = 5,
    force_new: bool = True
) -> MTLExcelLoggerV2:
    """
    创建 MTL Excel Logger V2 (工厂函数)

    Args:
        excel_path: Excel 文件路径
        checkpoint_name: checkpoint 名称
        config_dict: 配置字典 (用于基本信息和摘要)
        raw_yaml_content: 完整的 YAML 文件原始文本内容 (优先使用)
        n_folds: Fold 总数
        force_new: 是否强制创建新文件

    Returns:
        logger: MTLExcelLoggerV2 实例
    """
    return MTLExcelLoggerV2(
        excel_path=excel_path,
        checkpoint_name=checkpoint_name,
        config_dict=config_dict,
        raw_yaml_content=raw_yaml_content,
        n_folds=n_folds,
        force_new=force_new
    )


# =============================================================================
# 使用示例
# =============================================================================
"""
完整使用示例：

# 1. 初始化并写入 config_snapshot
config_dict = {
    "tasks": {
        "t1": {"name": "运动心功能分级", "num_classes": 3, "branch": "alpha"},
        "t2": {"name": "运动耐量", "num_classes": 3, "branch": "beta"},
        "t3": {"name": "标准心电运动负荷试验", "num_classes": 2, "branch": "beta"},
        ...
    },
    "training": {"epochs": 100, "batch_size": 16, "lr": 0.0003},
    "data": {"test_ratio": 0.2, "n_folds": 5},
    "mtl": {"architecture": {"variant": "protected_dual_engine_asymmetric_v3"}, ...}
}

logger = create_mtl_excel_logger_v2(
    excel_path="results/mtl_eval_mtl_v31_t6_protected.xlsx",
    checkpoint_name="mtl_v31_t6_protected",
    config_dict=config_dict,
    n_folds=5,
    force_new=True
)

# 2. 每个 Fold 完成后写入 3 个 sheet
for fold in range(1, 6):
    # 验证集指标（使用最佳阈值计算）
    val_metrics = {
        "t1": {"acc": 0.85, "precision": 0.83, "recall": 0.87, "f1": 0.85, "macro_f1": 0.85},
        "t2": {"acc": 0.78, ...},
        "t3": {"acc": 0.72, "precision": 0.68, "recall": 0.75, "f1": 0.71, "macro_f1": 0.71, "auc": 0.82},
        ...
    }

    # Holdout 测试集指标（使用相同的最佳阈值）
    holdout_metrics = {
        "t1": {"acc": 0.83, ...},
        ...
    }

    # 阈值搜索结果（完整信息）
    thresholds = {
        "t3": {
            "best_threshold": 0.325,
            "best_f1": 0.7234,
            "baseline_f1": 0.6812,
            "improvement": 0.0422
        },
        "t4": {"best_threshold": 0.410, ...},
        "t5": {"best_threshold": 0.285, ...}
    }

    task_names = {
        "t1": "运动心功能分级",
        "t2": "运动耐量",
        "t3": "标准心电运动负荷试验",
        ...
    }

    # 一行调用写入 3 个 sheet
    logger.write_fold_results(fold, val_metrics, holdout_metrics, thresholds, task_names)

# 3. 五折完成后写入汇总
logger.write_summary_sheets(task_names)

# 最终 Excel 文件包含 19 个 sheet：
# - val_fold1 ~ val_fold5 (5 个)
# - holdout_fold1 ~ holdout_fold5 (5 个)
# - threshold_fold1 ~ threshold_fold5 (5 个)
# - val_summary, holdout_summary, threshold_summary, config_snapshot (4 个)
"""