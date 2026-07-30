#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
MTL Excel Logger - 多任务学习评估结果落盘模块
==============================================

功能：
1. 记录验证集五折结果（sheet1: val_cv_5fold）
2. 记录 Holdout 测试集结果（sheet2: holdout_test）
3. 记录实验配置快照（sheet3: config_snapshot）

特点：
- config_snapshot 在实验开始时立即写入并保存
- 每个 fold 完成后立即写入并保存（防止崩溃丢失）
- 五折完成后追加 mean/std 汇总
- 支持未来扩展到 checkpoint B

使用示例：
    # 初始化并写入 config_snapshot
    logger = MTLExcelLogger(
        excel_path="results/mtl_eval_checkpointA.xlsx",
        checkpoint_name="mtl_v31_t6_protected",
        config_dict={
            "tasks": {"t1": {...}, ...},
            "training": {"epochs": 100, ...},
            "data": {...}
        }
    )

    # 每个 fold 完成后写入 val 和 holdout 结果
    logger.append_val_fold_metrics(fold=1, task_metrics=val_metrics, thresholds=thresholds)
    logger.append_holdout_fold_metrics(fold=1, task_metrics=test_metrics, thresholds=thresholds)

    # 五折完成后追加汇总
    logger.append_summary_rows("val_cv_5fold")
    logger.append_summary_rows("holdout_test")

创建日期: 2026-04-22
"""

import os
from typing import Dict, Any, Optional, List
from datetime import datetime
import json

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[Warning] openpyxl 未安装，Excel 功能不可用")


class MTLExcelLogger:
    """
    MTL Excel Logger - 多任务学习评估结果落盘

    Excel 文件结构：
    - sheet1 (val_cv_5fold): 验证集五折结果 + mean/std 汇总
    - sheet2 (holdout_test): Holdout 测试集结果 + mean/std 汇总
    - sheet3 (config_snapshot): 实验配置快照

    列结构：
    - fold: Fold 编号
    - task_key: 任务键 (t1~t6)
    - task_name: 任务名称
    - acc: 准确率
    - precision: 精确率
    - recall: 召回率
    - f1: F1分数
    - macro_f1: Macro F1分数
    - auc: AUC (仅二分类任务)
    - best_threshold: 最佳阈值 (仅二分类任务)
    """

    # 标准列名
    COLUMNS = [
        "fold", "task_key", "task_name", "acc", "precision",
        "recall", "f1", "macro_f1", "auc", "best_threshold"
    ]

    # 汇总行标识
    SUMMARY_LABELS = ["MEAN", "STD", "MIN", "MAX"]

    def __init__(
        self,
        excel_path: str,
        checkpoint_name: str = "mtl_v31_t6_protected",
        config_dict: Optional[Dict[str, Any]] = None,
        force_new: bool = False
    ):
        """
        初始化 Excel Logger

        Args:
            excel_path: Excel 文件路径
            checkpoint_name: checkpoint 名称 (用于标识，默认 "mtl_v31_t6_protected")
            config_dict: 配置字典 (用于写入 config_snapshot)
            force_new: 是否强制创建新文件 (默认 False，追加已有文件)
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl 未安装，请先安装: pip install openpyxl")

        self.excel_path = excel_path
        self.checkpoint_name = checkpoint_name
        self.config_dict = config_dict or {}
        self.fold_count = 0

        # 样式定义
        self._setup_styles()

        # 初始化 Excel 文件
        self._init_excel(force_new)

        # 立即写入 config_snapshot 并保存
        if config_dict:
            self.write_config_snapshot(config_dict)
            self.save()
            print(f"[MTLExcelLogger] config_snapshot 已写入并保存")

    def _setup_styles(self):
        """设置 Excel 样式"""
        # 表头样式
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        # 普通单元格样式
        self.cell_alignment = Alignment(horizontal="center", vertical="center")
        self.number_format = "0.0000"

        # 汇总行样式
        self.summary_font = Font(bold=True, size=11, color="0000FF")
        self.summary_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

        # 边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.cell_border = thin_border

    def _init_excel(self, force_new: bool):
        """初始化 Excel 文件"""
        if force_new or not os.path.exists(self.excel_path):
            # 创建新文件
            self.wb = Workbook()

            # 创建三个 sheet
            self._create_sheet("val_cv_5fold")
            self._create_sheet("holdout_test")
            self._create_sheet("config_snapshot")

            # 删除默认 sheet (Sheet)
            if "Sheet" in self.wb.sheetnames:
                del self.wb["Sheet"]

            # 立即保存
            os.makedirs(os.path.dirname(self.excel_path), exist_ok=True)
            self.wb.save(self.excel_path)
            print(f"[MTLExcelLogger] 创建新文件: {self.excel_path}")
        else:
            # 加载已有文件
            self.wb = openpyxl.load_workbook(self.excel_path)
            print(f"[MTLExcelLogger] 加载已有文件: {self.excel_path}")

            # 确保 sheet 存在
            for sheet_name in ["val_cv_5fold", "holdout_test", "config_snapshot"]:
                if sheet_name not in self.wb.sheetnames:
                    self._create_sheet(sheet_name)

    def _create_sheet(self, sheet_name: str):
        """创建 sheet 并写入表头"""
        ws = self.wb.create_sheet(title=sheet_name)

        if sheet_name in ["val_cv_5fold", "holdout_test"]:
            # 写入表头
            for col_idx, col_name in enumerate(self.COLUMNS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
                cell.border = self.cell_border

            # 固定列宽
            ws.column_dimensions['A'].width = 8   # fold
            ws.column_dimensions['B'].width = 10  # task_key
            ws.column_dimensions['C'].width = 20  # task_name
            ws.column_dimensions['D'].width = 10  # acc
            ws.column_dimensions['E'].width = 12  # precision
            ws.column_dimensions['F'].width = 10  # recall
            ws.column_dimensions['G'].width = 10  # f1
            ws.column_dimensions['H'].width = 12  # macro_f1
            ws.column_dimensions['I'].width = 10  # auc
            ws.column_dimensions['J'].width = 15  # best_threshold

        elif sheet_name == "config_snapshot":
            # config_snapshot sheet 结构
            ws.column_dimensions['A'].width = 25  # config_key
            ws.column_dimensions['B'].width = 60  # config_value

    def write_config_snapshot(self, config_dict: Dict[str, Any]):
        """
        写入配置快照

        Args:
            config_dict: 配置字典
        """
        ws = self.wb["config_snapshot"]

        # 清空已有内容（保留表头）
        ws.delete_rows(2, ws.max_row)

        # 写入配置信息
        config_items = [
            ("checkpoint_name", self.checkpoint_name),
            ("created_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("excel_path", self.excel_path),
            ("---", "---"),  # 分隔线
        ]

        # 添加任务配置
        if "tasks" in config_dict:
            tasks_config = config_dict["tasks"]
            for task_key, task_cfg in tasks_config.items():
                config_items.append(
                    (f"task.{task_key}", json.dumps(task_cfg, ensure_ascii=False))
                )

        # 添加训练配置
        if "training" in config_dict:
            training_config = config_dict["training"]
            config_items.append(("---", "---"))
            for key, value in training_config.items():
                config_items.append((f"training.{key}", str(value)))

        # 添加数据配置
        if "data" in config_dict:
            data_config = config_dict["data"]
            config_items.append(("---", "---"))
            for key, value in data_config.items():
                config_items.append((f"data.{key}", str(value)))

        # 写入到 sheet
        for row_idx, (key, value) in enumerate(config_items, start=2):
            ws.cell(row=row_idx, column=1, value=key)
            ws.cell(row=row_idx, column=2, value=value)

    def append_val_fold_metrics(
        self,
        fold: int,
        task_metrics: Dict[str, Dict[str, float]],
        thresholds: Optional[Dict[str, float]] = None,
        task_names: Optional[Dict[str, str]] = None
    ):
        """
        追加验证集 fold 指标

        Args:
            fold: Fold 编号
            task_metrics: 任务指标字典 {"t1": {"acc": ..., "f1": ..., ...}, ...}
            thresholds: 阈值字典 {"t3": 0.325, ...} (仅二分类任务)
            task_names: 任务名称字典 {"t1": "运动心功能分级", ...}
        """
        ws = self.wb["val_cv_5fold"]

        # 找到当前最后一行
        max_row = ws.max_row

        # 写入每个任务的指标
        for task_key in sorted(task_metrics.keys()):
            metrics = task_metrics[task_key]
            row_idx = max_row + 1

            # 写入数据
            self._write_task_row(
                ws, row_idx, fold, task_key, metrics,
                thresholds=thresholds,
                task_names=task_names
            )

        # 立即保存
        self.save()
        self.fold_count += 1
        print(f"[MTLExcelLogger] Fold {fold} val 指标已写入并保存")

    def append_holdout_fold_metrics(
        self,
        fold: int,
        task_metrics: Dict[str, Dict[str, float]],
        thresholds: Optional[Dict[str, float]] = None,
        task_names: Optional[Dict[str, str]] = None
    ):
        """
        追加 Holdout 测试集 fold 指标

        Args:
            fold: Fold 编号
            task_metrics: 任务指标字典 {"t1": {"acc": ..., "f1": ..., ...}, ...}
            thresholds: 阈值字典 {"t3": 0.325, ...} (仅二分类任务)
            task_names: 任务名称字典 {"t1": "运动心功能分级", ...}
        """
        ws = self.wb["holdout_test"]

        # 找到当前最后一行
        max_row = ws.max_row

        # 写入每个任务的指标
        for task_key in sorted(task_metrics.keys()):
            metrics = task_metrics[task_key]
            row_idx = max_row + 1

            # 写入数据
            self._write_task_row(
                ws, row_idx, fold, task_key, metrics,
                thresholds=thresholds,
                task_names=task_names
            )

        # 立即保存
        self.save()
        print(f"[MTLExcelLogger] Fold {fold} holdout test 指标已写入并保存")

    def _write_task_row(
        self,
        ws,
        row_idx: int,
        fold: int,
        task_key: str,
        metrics: Dict[str, float],
        thresholds: Optional[Dict[str, float]] = None,
        task_names: Optional[Dict[str, str]] = None
    ):
        """
        写入单行任务指标

        Args:
            ws: worksheet
            row_idx: 行号
            fold: Fold 编号
            task_key: 任务键
            metrics: 指标字典
            thresholds: 阈值字典
            task_names: 任务名称字典
        """
        task_name = task_names.get(task_key, task_key) if task_names else task_key

        # 准备数据
        data = [
            fold,
            task_key,
            task_name,
            metrics.get("acc", metrics.get("accuracy", 0)),
            metrics.get("precision", 0),
            metrics.get("recall", 0),
            metrics.get("f1", 0),
            metrics.get("macro_f1", metrics.get("f1", 0)),
            metrics.get("auc", 0),
        ]

        # 阈值（仅二分类任务）
        if thresholds and task_key in thresholds:
            data.append(thresholds[task_key])
        else:
            data.append("N/A")

        # 写入单元格
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # 数值格式化
            if isinstance(value, float) and col_idx >= 4:
                cell.number_format = self.number_format
            elif isinstance(value, float) and col_idx == 10:
                cell.number_format = "0.000"

            cell.alignment = self.cell_alignment
            cell.border = self.cell_border

    def append_summary_rows(
        self,
        sheet_name: str,
        task_keys: Optional[List[str]] = None
    ):
        """
        追加 mean/std 汇总行

        Args:
            sheet_name: sheet 名称 ("val_cv_5fold" 或 "holdout_test")
            task_keys: 需要汇总的任务列表 (默认全部任务)
        """
        ws = self.wb[sheet_name]

        # 获取已有数据
        data_rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = []
            for col_idx in range(1, len(self.COLUMNS) + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(value)
            data_rows.append(row_data)

        if not data_rows:
            print(f"[MTLExcelLogger] {sheet_name} 无数据，跳过汇总")
            return

        # 按任务分组
        task_data = {}
        for row in data_rows:
            task_key = row[1]  # task_key 列
            if task_keys and task_key not in task_keys:
                continue
            if task_key not in task_data:
                task_data[task_key] = []
            task_data[task_key].append(row)

        # 计算汇总
        summary_rows = []
        for task_key in sorted(task_data.keys()):
            task_rows = task_data[task_key]

            # 提取数值列 (acc, precision, recall, f1, macro_f1, auc)
            numeric_cols = [3, 4, 5, 6, 7, 8]  # 列索引 (从 0 开始)

            # 计算各统计量
            for stat_label in self.SUMMARY_LABELS:
                summary_row = [stat_label, task_key, f"{task_key}_{stat_label}"]

                for col_idx in numeric_cols:
                    values = [row[col_idx] for row in task_rows if isinstance(row[col_idx], (int, float))]

                    if not values:
                        summary_row.append(0)
                        continue

                    if stat_label == "MEAN":
                        summary_row.append(sum(values) / len(values))
                    elif stat_label == "STD":
                        import statistics
                        summary_row.append(statistics.stdev(values) if len(values) > 1 else 0)
                    elif stat_label == "MIN":
                        summary_row.append(min(values))
                    elif stat_label == "MAX":
                        summary_row.append(max(values))

                # 阈值列 (汇总行不需要)
                summary_row.append("N/A")

                summary_rows.append(summary_row)

        # 写入汇总行
        max_row = ws.max_row + 2  # 空一行

        for row_data in summary_rows:
            row_idx = max_row
            max_row += 1

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # 汇总行样式
                cell.font = self.summary_font
                cell.fill = self.summary_fill
                cell.alignment = self.cell_alignment
                cell.border = self.cell_border

                # 数值格式化
                if isinstance(value, float) and col_idx >= 4:
                    cell.number_format = self.number_format

        # 立即保存
        self.save()
        print(f"[MTLExcelLogger] {sheet_name} mean/std 汇总已写入并保存")

    def save(self):
        """强制保存 Excel 文件"""
        self.wb.save(self.excel_path)


def create_mtl_excel_logger(
    excel_path: str,
    checkpoint_name: str = "mtl_v31_t6_protected",
    config_dict: Optional[Dict[str, Any]] = None,
    force_new: bool = False
) -> MTLExcelLogger:
    """
    创建 MTL Excel Logger (工厂函数)

    Args:
        excel_path: Excel 文件路径
        checkpoint_name: checkpoint 名称
        config_dict: 配置字典
        force_new: 是否强制创建新文件

    Returns:
        logger: MTLExcelLogger 实例
    """
    return MTLExcelLogger(
        excel_path=excel_path,
        checkpoint_name=checkpoint_name,
        config_dict=config_dict,
        force_new=force_new
    )


# =============================================================================
# 使用示例 (完整流程)
# =============================================================================
"""
完整使用示例：

# 1. 初始化并写入 config_snapshot
config_dict = {
    "tasks": {
        "t1": {"name": "运动心功能分级", "num_classes": 3, "branch": "alpha"},
        "t2": {"name": "运动耐量", "num_classes": 3, "branch": "beta"},
        ...
    },
    "training": {
        "epochs": 100,
        "batch_size": 16,
        "lr": 0.0003
    },
    "data": {
        "test_ratio": 0.2,
        "n_folds": 5
    }
}

logger = create_mtl_excel_logger(
    excel_path="results/mtl_eval_checkpointA.xlsx",
    checkpoint_name="mtl_v31_t6_protected",
    config_dict=config_dict,
    force_new=True
)

# 2. 每个 Fold 完成后写入结果
for fold in range(1, 6):
    # 验证集评估（阈值搜索后）
    val_metrics = {
        "t1": {"acc": 0.85, "precision": 0.83, "recall": 0.87, "f1": 0.85, "macro_f1": 0.85},
        "t2": {...},
        "t3": {"acc": 0.72, "precision": 0.68, "recall": 0.75, "f1": 0.71, "macro_f1": 0.71, "auc": 0.82},
        ...
    }
    thresholds = {"t3": 0.325, "t4": 0.410, "t5": 0.285}

    logger.append_val_fold_metrics(fold, val_metrics, thresholds, task_names)

    # Holdout 测试集评估
    test_metrics = {
        "t1": {"acc": 0.83, "precision": 0.80, "recall": 0.85, "f1": 0.82, "macro_f1": 0.82},
        ...
    }

    logger.append_holdout_fold_metrics(fold, test_metrics, thresholds, task_names)

# 3. 五折完成后追加汇总
logger.append_summary_rows("val_cv_5fold")
logger.append_summary_rows("holdout_test")
"""